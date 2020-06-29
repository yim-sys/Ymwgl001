from openpyxl import load_workbook


class ExcelHandler:
    def __init__(self,file_path):
        """初始化路径"""
        self.file_path=file_path
        workbook=None

    def get_sheet(self,sheetname):
        """获取工作表前打开工作薄"""
        workbook=load_workbook(self.file_path)
        sheet=workbook[sheetname]
        return sheet

    def read_data(self,sheetname):
        """读取数据"""
        sheet=self.get_sheet(sheetname)
        #获取所有行数
        rows = list(sheet.rows)
        #获取标题
        titles=[]
        for title in rows[0]:
            titles.append(title.value)
        #存取所有行的数据
        datas=[]
        for row in rows[1:]:
            case={} #存取一行的数据
            for idx,cell in enumerate (row):
                case[titles[idx]]=cell.value
            datas.append(case)
        return datas

    def write(self,file_path=None,sheet_name=None,row=0,col=0,data=None):
        workbook=load_workbook(file_path)
        sheet=workbook[sheet_name]
        sheet.cell(row,col).value=data
        workbook.save(file_path)
        workbook.close()


if __name__ == '__main__':
    b=ExcelHandler(r'D:\python_t\futureloanFrame\data\cases.xlsx')
    case_data=b.read_data('register')
    print(case_data)
    w=b.write(r'D:\python_t\futureloanFrame\data\cases.xlsx','register', 5, 9, '通过')
